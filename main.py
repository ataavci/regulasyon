import numpy as np
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
import os



def dosya_erisim(path, sheet_name, col):
    
    workbook =load_workbook(filename=path)
    if sheet_name not in workbook.sheetnames:
        print("HATALI SHEET NAME")
        return
    sheet = workbook[sheet_name]
    for cell in sheet[col]:
        print(f"{cell.column_letter}{cell.row}={cell.value}")

if __name__ == "__main__":
    dosya_erisim("ets_hesap.xlsx", sheet_name="ETS CONTENT", col=["A","B"])